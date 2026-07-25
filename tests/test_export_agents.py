"""Excel export, enrichment merge, and cheap-model routing."""

from __future__ import annotations

from openpyxl import load_workbook

from job_hunter import enrich, export
from job_hunter.agents import Complexity, pick_model


def test_pick_model_is_cheap_first():
    assert "haiku" in pick_model(Complexity.trivial)
    assert "haiku" in pick_model("simple")
    assert "sonnet" in pick_model(Complexity.complex)


def test_apply_enrichment_sets_fields(make_job):
    job = make_job("Backend Engineer")
    enrich.apply_enrichment(job, {
        "about": "Makes widgets.", "salary": "$120k-$150k/yr (USD)",
        "qualifications": "3+ yrs Python", "work_culture": "Fast-paced, remote-friendly.",
        "pros": "Good pay; flexible", "cons": "Long hours", "source": "Glassdoor",
    })
    assert job.enriched is True
    assert job.salary == "$120k-$150k/yr (USD)"
    assert job.work_culture and "remote-friendly" in job.work_culture
    assert job.pros and job.cons


def test_enrichment_prompt_mentions_json(make_job):
    prompt = enrich.enrichment_prompt(make_job("Data Scientist", company="Acme"))
    assert "Acme" in prompt and "JSON" in prompt


def test_to_excel_writes_rows_and_new_columns(make_job, tmp_path):
    jobs = [make_job("Backend Engineer"), make_job("Data Scientist", company="Beta")]
    jobs[0].salary = "₹18-24 LPA (INR)"
    jobs[0].posted_ago = "2 weeks ago"
    jobs[0].num_applicants = "47 applicants"
    jobs[0].work_culture = "Collaborative"
    out = export.to_excel(jobs, tmp_path / "jobs.xlsx")
    wb = load_workbook(out)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    for h in ["Posted", "Applicants", "Work culture", "Pros (reviews)",
              "Cons (reviews)", "Job summary"]:
        assert h in headers
    assert ws.max_row == 3
    values = [c.value for c in ws[2]]
    assert "2 weeks ago" in values and "₹18-24 LPA (INR)" in values


def test_summary_is_truncated(make_job, tmp_path):
    job = make_job("Engineer")
    job.jd_summary = "x" * 2000
    out = export.to_excel([job], tmp_path / "j.xlsx")
    ws = load_workbook(out).active
    col = [c.value for c in ws[1]].index("Job summary") + 1
    assert len(ws.cell(row=2, column=col).value) <= 701   # 700 + ellipsis
