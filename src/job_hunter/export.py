"""Export jobs to an Excel workbook the user can browse and pick from."""

from __future__ import annotations

from pathlib import Path

from .models import Job

_COLUMNS = [
    ("id", "ID", 14),
    ("title", "Title", 32),
    ("company", "Company", 22),
    ("location", "Location", 20),
    ("workplace_type", "Workplace", 11),
    ("posted_ago", "Posted", 12),
    ("num_applicants", "Applicants", 16),
    ("rag", "Eligibility", 12),
    ("flags", "Flags / concerns (why R/Y)", 44),
    ("salary", "Salary (avg, w/ currency)", 26),
    ("about", "About the company", 40),
    ("work_culture", "Work culture", 36),
    ("pros", "Pros (reviews)", 36),
    ("cons", "Cons (reviews)", 36),
    ("qualifications", "Qualifications", 40),
    ("description", "Job description", 60),
    ("status", "Applied?", 11),
    ("url", "Link", 38),
    ("enrichment_source", "Source", 26),
]

# Columns to truncate so the sheet stays readable.
_TRUNCATE = {"description": 800}

# RAG cell colors.
_RAG_FILL = {"green": "C6EFCE", "yellow": "FFEB9C", "red": "FFC7CE"}


def to_excel(jobs: list[Job], path: str | Path | None = None) -> str:
    """Write jobs to an .xlsx and return the path."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    # Default to the current working directory (i.e. the repo folder when run
    # from there) so the spreadsheet is easy to find and open — not buried in
    # ~/.jobhunter with the internal state.
    out = Path(path) if path else Path.cwd() / "jobs.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, (_, header, width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[cell.column_letter].width = width
    ws.freeze_panes = "A2"

    wrap = Alignment(wrap_text=True, vertical="top")
    for row_idx, job in enumerate(jobs, start=2):
        rag = getattr(job, "rag", None)
        rag_fill = PatternFill("solid", fgColor=_RAG_FILL[rag]) if rag in _RAG_FILL else None
        for col_idx, (attr, _, _) in enumerate(_COLUMNS, start=1):
            value = getattr(job, attr, None)
            value = value.value if hasattr(value, "value") else value  # StrEnum
            if attr == "rag":
                value = (value or "").upper()
            if attr in _TRUNCATE and isinstance(value, str) and len(value) > _TRUNCATE[attr]:
                value = value[: _TRUNCATE[attr]] + "…"
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = wrap
            # Colour the Eligibility + Flags cells by RAG so they pop.
            if attr in ("rag", "flags") and rag_fill is not None:
                cell.fill = rag_fill

    wb.save(out)
    return str(out)
